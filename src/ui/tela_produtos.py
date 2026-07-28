import tkinter as tk

from tkinter import ttk, messagebox, simpledialog
from tkinter import filedialog
from wsgiref import headers

from openpyxl import styles

from src.modules.produtos import listar_produtos, inserir_produto, atualizar_produto, excluir_produto, salvar_produto
from src.modules.estoque import entrada_estoque, saida_estoque
from src.modules.db_config import obter_config
from src.ui import styles
from src.utils.cores import *
from src.ui.styles import *
from src.utils.botoes import botao_moderno, botao_menor
from src.utils.formatacao import agora_brasil, moeda

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, PatternFill

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4

class TelaProdutos(tk.Frame):

    def __init__(self, master, tema):
        super().__init__(master)
        
        style = ttk.Style()
        style.theme_use("default")
        self.tema = tema

        style.configure(
            "Treeview",
            rowheight=28,
            font=self.tema["font"]
        )
        style.map(
            "Treeview",
            background=[("selected", "#22c55e")],
            foreground=[("selected", "white")]
        )
        style.configure(
            "Treeview",
            bordercolor=self.tema["primary"],
            relief="flat"
        )
        style.configure(
            "Treeview.Heading",
            font=self.tema["title"]
        )

        self.pack(fill="both", expand=True)
        
        self.configure(bg=self.tema["card"])

        # ===============================
        # BARRA SUPERIOR
        # ===============================

        barra_topo = tk.Frame(self)
        barra_topo.pack(fill="x", padx=10, pady=5)

        # ===============================
        # BARRA DE BUSCA
        # ===============================

        barra_topo = tk.Frame(self, bg=self.tema["card"])
        barra_topo.pack(fill="x", padx=10, pady=5)
        barra_topo.configure(height=50)

        # TÍTULO
        titulo = tk.Label(
            barra_topo,
            text="Produtos",
            fg=self.tema["text"],
            bg=self.tema["card"],
            font=self.tema["title"]
        )
        titulo.pack(side="left", padx=(0, 20))

        # BUSCA
        tk.Label(
            barra_topo,
            bg=self.tema["card"],
            fg=self.tema["text"]
        ).pack(side="left")

        self.entry_busca = tk.Entry(
            barra_topo,
            bg="#ffffff",
            fg="black",
            insertbackground="white",
            width=25
        )
        self.entry_busca.pack(side="left", padx=5)

        btn_buscar = botao_menor(
            barra_topo,
            "Buscar 🔍",
            self.buscar_produto,
            "default"
        )
        btn_buscar.pack(side="left", padx=5)

        # BOTÃO NOVO (fica na direita)
        btn_novo = botao_moderno(barra_topo, "+ Novo Produto", self.novo_produto, "primary")
        btn_novo.pack(side="right", padx=5)

        # ==========================
        # AREA PRINCIPAL
        # ==========================

        frame_conteudo = tk.Frame(self)
        frame_conteudo.pack(fill="both", expand=True)

        # ==========================
        # LISTA PRODUTOS
        # ==========================

        frame_lista = tk.Frame(frame_conteudo, width=700)
        frame_lista.pack(side="left", fill="both", padx=10, pady=10)
        frame_lista.pack_propagate(False)

        self.tabela = ttk.Treeview(
            frame_lista,
            columns=("id","codigo","nome","preco","estoque"),
            show="headings"
        )
        
        self.tabela.tag_configure(
            "estoque_baixo",
            background="#c45d67",
            foreground="white"
        )
        style.map(
            "Treeview",
            background=[("selected", "#5a90c2")],
            foreground=[("selected", "white")]
        )

        self.tabela.heading("id", text="ID")
        self.tabela.heading("codigo", text="Código")
        self.tabela.heading("nome", text="Produto")
        self.tabela.heading("preco", text="Preço")
        self.tabela.heading("estoque", text="Estoque")

        self.tabela.column("id", width=50)
        self.tabela.column("codigo", width=120)
        self.tabela.column("nome", width=250)
        self.tabela.column("preco", width=100)
        self.tabela.column("estoque", width=100)

        self.tabela.pack(fill="both", expand=True)
        self.tabela.tag_configure(
            "estoque_baixo",
            background="#c45d67",
            foreground="white"
        )

        self.tabela.bind("<<TreeviewSelect>>", self.mostrar_detalhes)

        self.pagina_atual = 0
        self.limite = 20
        self.total_registros = 0
        self.filtro_busca = ""

        frame_paginacao = tk.Frame(frame_lista)
        frame_paginacao.pack(fill="x", pady=(5, 0))

        self.btn_anterior = botao_menor(
            frame_paginacao,
            "Anterior",
            self.pagina_anterior,
            "default"
        )
        self.btn_anterior.pack(side="left", padx=5)

        self.label_pagina = tk.Label(
            frame_paginacao,
            text="Página 1 de 1",
            font=("Arial", 10, "bold")
        )
        self.label_pagina.pack(side="left", padx=10)

        self.btn_proximo = botao_menor(
            frame_paginacao,
            "Próxima",
            self.proxima_pagina,
            "default"
        )
        self.btn_proximo.pack(side="left", padx=5)

        # ==========================
        # DETALHES
        # ==========================

        frame_detalhes = tk.Frame(
            frame_conteudo,
            width=350,
            bg=self.tema["card"]
        )
        frame_detalhes.pack(side="right", fill="y", padx=10, pady=10)
        frame_detalhes.pack_propagate(False)

        titulo_det = tk.Label(
            frame_detalhes,
            text="Detalhes do Produto",
            font=("Arial", 14, "bold"),
            bg=self.tema["card"]
        )

        titulo_det.pack(pady=10)

        self.label_info = tk.Label(
            frame_detalhes,
            text="Selecione um produto",
            justify="left",
            bg=self.tema["card"],
            font=self.tema["font"]
        )

        self.label_info.pack(padx=10, pady=10)

        # ==========================
        # BOTÕES
        # ==========================

        frame_acoes = tk.Frame(frame_detalhes, bg=self.tema["card"])
        frame_acoes.pack(pady=20, fill="x")

        btn_entrada = botao_moderno(frame_acoes, "Entrada", self.entrada_estoque, "primary")
        btn_entrada.pack(fill="x", pady=5)

        btn_saida = botao_moderno(frame_acoes, "Saída", self.saida_estoque, "warning")
        btn_saida.pack(fill="x", pady=5)
        
        btn_editar = botao_moderno(frame_acoes, "Editar", self.editar_produto,"primary")
        btn_editar.pack(fill="x", pady=5)

        btn_excluir = botao_moderno(frame_acoes, "Excluir", self.excluir_produto, "danger")
        btn_excluir.pack(fill="x", pady=5)
        
        btn_exportar = botao_menor(frame_detalhes, "📥 Exportar Excel", self.exportar_excel, "default")
        btn_exportar.pack(fill="x", pady=5)
        
        btn_exportar_pdf = botao_menor(frame_detalhes, "📄 Exportar PDF", self.exportar_pdf, "default")
        btn_exportar_pdf.pack(fill="x", pady=5)

        # CARREGA PRODUTOS
        self.carregar_produtos()

    def carregar_produtos(self):
        
        estoque_minimo_global = obter_config("estoque_minimo", 5)
        estoque_minimo_global = int(estoque_minimo_global)

        for item in self.tabela.get_children():
            self.tabela.delete(item)

        produtos = listar_produtos()

        if self.filtro_busca:
            termo = self.filtro_busca.lower()
            produtos = [
                p for p in produtos
                if termo in (p[2] or "").lower() or termo in (p[1] or "").lower()
            ]

        self.total_registros = len(produtos)

        inicio = self.pagina_atual * self.limite
        fim = inicio + self.limite
        pagina_atual = produtos[inicio:fim]

        for p in pagina_atual:

            id_produto = p[0]
            codigo = p[1]
            nome = p[2]
            preco = p[3]
            estoque = p[4]
            estoque_minimo = p[5] if len(p) > 5 else 5

            if estoque is None:
                estoque = 0

            if estoque_minimo is None:
                estoque_minimo = estoque_minimo_global

            alerta_ativo = obter_config("alerta_estoque", "1")

            if alerta_ativo == "1" and estoque <= estoque_minimo:

                self.tabela.insert(
                    "",
                    "end",
                    values=(id_produto, codigo, nome, moeda(preco), estoque),
                    tags=("estoque_baixo",)
                )

            else:

                self.tabela.insert(
                    "",
                    "end",
                    values=(id_produto, codigo, nome, moeda(preco), estoque)
                )

        self.atualizar_label_pagina()
        self.atualizar_botoes_paginacao()
            
    def buscar_produto(self):

        self.filtro_busca = self.entry_busca.get().strip()
        self.pagina_atual = 0
        self.carregar_produtos()

    def proxima_pagina(self):
        total_paginas = max(1, (self.total_registros + self.limite - 1) // self.limite)
        if self.pagina_atual < total_paginas - 1:
            self.pagina_atual += 1
            self.carregar_produtos()

    def pagina_anterior(self):
        if self.pagina_atual > 0:
            self.pagina_atual -= 1
            self.carregar_produtos()

    def atualizar_label_pagina(self):
        total_paginas = max(1, (self.total_registros + self.limite - 1) // self.limite)
        self.label_pagina.config(text=f"Página {self.pagina_atual + 1} de {total_paginas}")

    def atualizar_botoes_paginacao(self):
        total_paginas = max(1, (self.total_registros + self.limite - 1) // self.limite)

        if self.pagina_atual == 0:
            self.btn_anterior.config(state="disabled")
        else:
            self.btn_anterior.config(state="normal")

        if self.pagina_atual >= total_paginas - 1:
            self.btn_proximo.config(state="disabled")
        else:
            self.btn_proximo.config(state="normal")

    def obter_produtos_filtrados(self):
        produtos = listar_produtos()

        if self.filtro_busca:
            termo = self.filtro_busca.lower()
            produtos = [
                p for p in produtos
                if termo in (p[2] or "").lower() or termo in (p[1] or "").lower()
            ]

        return produtos

    def mostrar_detalhes(self, event):

        selecionado = self.tabela.selection()

        if not selecionado:
            return

        item = selecionado[0]
        dados = self.tabela.item(item)["values"]

        texto = f"""
            ID: {dados[0]}
            Código: {dados[1]}
            Produto: {dados[2]}

            Preço: R$ {dados[3]}
            Estoque: {dados[4]}
        """

        self.label_info.config(text=texto)

    def novo_produto(self):

        janela = tk.Toplevel(self)
        janela.title("Cadastro de Produto")
        janela.geometry("350x300")
        janela.grab_set()

        tk.Label(janela, text="Código").pack()
        entry_codigo = tk.Entry(janela)
        entry_codigo.pack()
        entry_codigo.focus()

        tk.Label(janela, text="Nome").pack()
        entry_nome = tk.Entry(janela)
        entry_nome.pack()

        tk.Label(janela, text="Preço").pack()
        entry_preco = tk.Entry(janela)
        entry_preco.pack()

        tk.Label(janela, text="Estoque").pack()
        entry_estoque = tk.Entry(janela)
        entry_estoque.pack()


    # ===============================
    # FUNÇÃO SALVAR
    # ===============================

        def salvar():

            codigo = entry_codigo.get().strip()
            nome = entry_nome.get().strip()
            preco_venda = entry_preco.get().strip()
            estoque = entry_estoque.get().strip()

            # VALIDAÇÃO
            if not codigo or not nome or not preco_venda or not estoque:
                messagebox.showerror(
                    "Erro",
                    "Todos os campos são obrigatórios."
                )
                return

        # VALIDAR NÚMEROS
            try:
                preco_venda = float(preco_venda)
                estoque = int(estoque)
            except ValueError:
                messagebox.showerror(
                    "Erro",
                    "Preço deve ser número e estoque deve ser inteiro."
                )
                return

        # SALVAR NO BANCO
            try:

                inserir_produto(codigo, nome, preco_venda, estoque)

                messagebox.showinfo(
                    "Sucesso",
                    "Produto cadastrado com sucesso!"
                )

                janela.destroy()

                self.carregar_produtos()

            except Exception as e:

                if "UNIQUE constraint" in str(e):

                    messagebox.showerror(
                        "Erro",
                        "Já existe um produto com esse código."
                    )

                else:

                    messagebox.showerror(
                        "Erro",
                        f"Erro ao salvar produto:\n{e}"
                    )

    # BOTÃO SALVAR
        btn_salvar = botao_moderno(
            janela,
            "Salvar",
            salvar,
            "primary"
        )

        btn_salvar.pack(pady=10)

    def editar_produto(self):

        selecionado = self.tabela.selection()

        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um produto.")
            return

        item = self.tabela.selection()[0]
        
        dados = self.tabela.item(item)["values"]

        id_produto = dados[0]

        janela = tk.Toplevel(self)
        janela.title("Editar Produto")
        janela.geometry("350x300")
        janela.grab_set()

        tk.Label(janela, text="Código").pack()
        entry_codigo = tk.Entry(janela)
        entry_codigo.pack()
        entry_codigo.insert(0, dados[1])

        tk.Label(janela, text="Nome").pack()
        entry_nome = tk.Entry(janela)
        entry_nome.pack()
        entry_nome.insert(0, dados[2])

        tk.Label(janela, text="Preço").pack()
        entry_preco = tk.Entry(janela)
        entry_preco.pack()
        entry_preco.insert(0, dados[3])

        tk.Label(janela, text="Estoque").pack()
        entry_estoque = tk.Entry(janela)
        entry_estoque.pack()
        entry_estoque.insert(0, dados[4])

        def salvar():

            codigo = entry_codigo.get().strip()
            nome = entry_nome.get().strip()
            preco = entry_preco.get().strip()
            estoque = entry_estoque.get().strip()

            if not codigo or not nome or not preco or not estoque:
                messagebox.showerror("Erro", "Todos os campos são obrigatórios.")
                return

            try:
                preco = preco.strip()
                if preco.startswith("R$"):
                    preco = preco[2:].strip()
                preco = float(preco.replace(".", "").replace(",", "."))
                estoque = int(estoque)
            except ValueError:
                messagebox.showerror("Erro", "Preço inválido ou estoque inválido.")
                return

            try:

                atualizar_produto(id_produto, codigo, nome, preco, estoque)

                messagebox.showinfo("Sucesso", "Produto atualizado!")

                janela.destroy()
                self.carregar_produtos()

            except Exception as e:

                messagebox.showerror("Erro", str(e))

        btn_salvar = botao_moderno(
            janela,
            "Salvar Alterações",
            salvar,
            "primary"
        )

        btn_salvar.pack(pady=10)
        
    def excluir_produto(self):

        selecionado = self.tabela.selection()

        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um produto.")
            return

        item = self.tabela.selection()[0]
        
        dados = self.tabela.item(item)["values"]

        id_produto = dados[0]
        nome_produto = dados[2]

        confirmar = messagebox.askyesno(
            "Confirmar exclusão",
            f"Deseja excluir o produto:\n\n{nome_produto} ?"
        )

        if not confirmar:
            return

        try:

            excluir_produto(id_produto)

            messagebox.showinfo(
                "Sucesso",
                "Produto excluído com sucesso!"
            )

            self.carregar_produtos()

            self.label_info.config(text="Selecione um produto")

        except Exception as e:

            messagebox.showerror(
                "Erro",
                f"Erro ao excluir produto:\n{e}"
            )
            
    def entrada_estoque(self):

        selecionado = self.tabela.selection()

        if not selecionado:
            return

        item = selecionado[0]
        dados = self.tabela.item(item)["values"]
        produto_id = dados[0]

        quantidade = simpledialog.askinteger(
            "Entrada de Estoque",
            "Quantidade a adicionar:",
            minvalue=1
        )

        if quantidade:

            entrada_estoque(produto_id, quantidade)

            messagebox.showinfo("Sucesso", "Estoque atualizado")

            self.carregar_produtos()
            
    def saida_estoque(self):

        selecionado = self.tabela.selection()

        if not selecionado:
            return

        item = selecionado[0]
        dados = self.tabela.item(item)["values"]
        produto_id = dados[0]

        quantidade = simpledialog.askinteger(
            "Saída de Estoque",
            "Quantidade a retirar:"
        )

        if quantidade:
            try:
                saida_estoque(produto_id, quantidade)
                messagebox.showinfo("Sucesso", "Saída registrada!")
                self.carregar_produtos()
            except Exception as e:
                messagebox.showerror("Erro", str(e))
            
    def exportar_excel(self):

        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx")],
            title="Salvar planilha"
        )

        if not caminho:
            return

        wb = Workbook()
        ws = wb.active

        ws.title = "Produtos"

        # CABEÇALHO
        ws["A1"] = "INTEGRIS ERP"
        ws["A2"] = "Relatório de Produtos"

        ws["A4"] = f"Gerado em: {agora_brasil().strftime('%d/%m/%Y %H:%M:%S')}"

        # TÍTULOS
        headers = ["Código", "Produto", "Estoque", "Preço"]

        linha_header = 6

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=linha_header, column=col)
            cell.value = header

            cell.font = Font(bold=True, color="FFFFFF")

            cell.fill = PatternFill(
                start_color="1F4E78",
                end_color="1F4E78",
                fill_type="solid"
            )

        # DADOS
        linha = 7

        total_qtd_produtos = 0

        produtos = self.obter_produtos_filtrados()

        for produto in produtos:
            codigo = produto[1]
            nome = produto[2]
            preco = produto[3]
            estoque = produto[4] or 0

            ws.cell(linha, 1, codigo)
            ws.cell(linha, 2, nome)
            ws.cell(linha, 3, estoque)
            ws.cell(linha, 4, moeda(preco))

            total_qtd_produtos += int(estoque)
            linha += 1

        total_produtos = len(produtos)

        ws.cell(linha + 2, 1, f"Total de produtos cadastrados: {total_produtos}")
        ws.cell(linha + 3, 1, f"Total de produtos em estoque: {total_qtd_produtos}")
        # LARGURA COLUNAS
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

        wb.save(caminho)

        messagebox.showinfo(
            "Exportação",
            "Planilha exportada com sucesso!"
        )
        
    def exportar_pdf(self):

        caminho = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Arquivo PDF", "*.pdf")],
            title="Salvar PDF"
        )

        if not caminho:
            return

        documento = SimpleDocTemplate(
            caminho,
            pagesize=A4
        )

        elementos = []

        styles = getSampleStyleSheet()

        # CABEÇALHO
        titulo = Paragraph(
            "<b>INTEGRIS ERP</b><br/>Relatório de Produtos",
            styles["Title"]
        )

        elementos.append(titulo)

        elementos.append(Spacer(1, 20))

        data = Paragraph(
            f"Gerado em: {agora_brasil().strftime('%d/%m/%Y %H:%M:%S')}",
            styles["Normal"]
        )

        elementos.append(data)

        elementos.append(Spacer(1, 20))

        # TABELA
        dados = [
            ["Código", "Produto", "Estoque", "Preço (Un.)"]
        ]

        for produto in self.obter_produtos_filtrados():
            codigo = produto[1]
            nome = produto[2]
            preco = produto[3]
            estoque = produto[4] or 0

            dados.append([
                codigo,
                nome,
                estoque,
                moeda(preco)
            ])

        tabela = Table(dados)

        tabela.setStyle(TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")),

            ("TEXTCOLOR", (0,0), (-1,0), colors.white),

            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

            ("GRID", (0,0), (-1,-1), 1, colors.black),

            ("BACKGROUND", (0,1), (-1,-1), colors.beige),

            ("BOTTOMPADDING", (0,0), (-1,0), 10),

        ]))

        elementos.append(tabela)

        elementos.append(Spacer(1, 20))

        total_produtos = len(self.obter_produtos_filtrados())

        total = Paragraph(
            f"<b>Total de produtos:</b> {total_produtos}",
            styles["Normal"]
        )

        elementos.append(total)

        documento.build(elementos)

        messagebox.showinfo(
            "PDF",
            "PDF exportado com sucesso!"
        )
            
    def criar_header(self, titulo_texto):

        header = tk.Frame(self, bg="white")
        header.pack(fill="x")

        titulo = tk.Label(
            header,
            text=titulo_texto,
            font=("Arial", 18, "bold"),
            bg="white"
        )
        titulo.pack(side="left", padx=10, pady=10)

        btn_novo = botao_moderno(
            header,
            "+ Novo Produto",
            self.novo_produto,
            "sucess"
        )
        btn_novo.pack(side="right")
        
        # ==========================
        # AREA PRINCIPAL
        # ==========================

        frame_conteudo = tk.Frame(self)
        frame_conteudo.pack(fill="both", expand=True)

        # ==========================
        # LISTA PRODUTOS
        # ==========================

        frame_lista = tk.Frame(frame_conteudo, width=700)
        frame_lista.pack(side="left", fill="both", padx=10, pady=10)
        frame_lista.pack_propagate(False)

        self.tabela = ttk.Treeview(
            frame_lista,
            columns=("id","codigo","nome","preco","estoque"),
            show="headings"
        )
        
        self.tabela.tag_configure(
            "estoque_baixo",
            background="#ff4d4d",
            foreground="white"
        )

        self.tabela.heading("id", text="ID")
        self.tabela.heading("codigo", text="Código")
        self.tabela.heading("nome", text="Produto")
        self.tabela.heading("preco", text="Preço")
        self.tabela.heading("estoque", text="Estoque")

        self.tabela.column("id", width=50)
        self.tabela.column("codigo", width=120)
        self.tabela.column("nome", width=250)
        self.tabela.column("preco", width=100)
        self.tabela.column("estoque", width=100)

        self.tabela.pack(fill="both", expand=True)
        self.tabela.tag_configure(
            "estoque_baixo",
            background="#ffcccc"
        )

        self.tabela.bind("<<TreeviewSelect>>", self.mostrar_detalhes)

        # ==========================
        # DETALHES
        # ==========================

        frame_detalhes = tk.Frame(frame_conteudo, width=350, bg="#ecf0f1")
        frame_detalhes.pack(side="right", fill="y", padx=10, pady=10)
        frame_detalhes.pack_propagate(False)

        titulo_det = tk.Label(
            frame_detalhes,
            text="Detalhes do Produto",
            font=("Arial", 14, "bold"),
            bg="#ecf0f1"
        )

        titulo_det.pack(pady=10)

        self.label_info = tk.Label(
            frame_detalhes,
            text="Selecione um produto",
            justify="left",
            bg=self.tema["bg"],
            font=self.tema["text"]
        )

        self.label_info.pack(padx=10, pady=10)

        # BOTÕES

        frame_acoes = tk.Frame(frame_detalhes, bg="#ecf0f1")
        frame_acoes.pack(pady=20)

        btn_editar = tk.Button(
            frame_acoes,
            text="Editar",
            command=self.editar_produto
        )
        btn_editar.pack(side="left", padx=5)

        btn_excluir = tk.Button(
            frame_acoes,
            text="Excluir",
            bg="#e74c3c",
            fg="white",
            command=self.excluir_produto,
            font=self.tema["text"]
        )
        btn_excluir.pack(side="left", padx=5)
        
        btn_entrada = tk.Button(
            frame_acoes,
            text="Entrada",
            bg="#2eaccc",
            fg="white",
            command=self.entrada_estoque
        )
        btn_entrada.pack(side="left", padx=5)

        btn_saida = tk.Button(
            frame_acoes,
            text="Saída",
            bg="#2eaccc",
            fg="white",
            command=self.saida_estoque
        )
        btn_saida.pack(side="left", padx=5)

        # CARREGA PRODUTOS
        self.carregar_produtos()